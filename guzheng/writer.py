# -*- coding: utf-8 -*-
"""
輸出工作簿組裝層。
將所有業務資料與換場任務寫入 Excel，並套用樣式。
"""
from __future__ import annotations
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .config import SONG_COLOR_MAP, INTERMISSION_LABEL, ENDING_BLOCK, INTERMISSION_AFTER
from .models import SongAsset, SongPeople
from .scheduler import generate_transition
from .styles import (
    HEADER_FONT_SIZE,
    header_fill, song_fill, intermission_fill, thin_border,
    style_cell, style_row,
)
from .utils import stand_for_performer

# key 型別別名，與 timeline.py 保持一致
TransitionKey = Tuple[Optional[str], Optional[str], str]

# 欄位定義：(欄號, 表頭文字, 寬度)
_COLUMNS = [
    (1, "表演時間",                16),
    (2, "曲目",                   20),
    (3, "時間長度/\n古箏定位顏色", 18),
    (4, "表演人員/古箏/箏架",      28),
    (5, "下箏/樂器（表演者自己來）", 26),
    (6, "箏架、譜架、椅子",        28),
    (7, "上箏",                   24),
    (8, "備註",                   18),
]
NUM_COLS = len(_COLUMNS)


# =============================================================
# 儲存格文字格式化
# =============================================================

def _format_performer_cell(song_people: SongPeople) -> str:
    """將 SongPeople 轉為表演人員欄的多行顯示文字。"""
    lines = []
    for name, guzheng, stand in song_people.guzheng_players:
        lines.append(f"【{name}】{guzheng}/{stand_for_performer(stand)}")
    for name in song_people.percussion_players:
        lines.append(f"【{name}】打擊")
    for name in song_people.piano_players:
        lines.append(f"【{name}】電鋼琴/立架")
    for name in song_people.bass_players:
        lines.append(f"【{name}】低音提琴")
    return "\n".join(lines)


def _duration_color_text(song: str, duration_map: Dict[str, str]) -> str:
    """組合「時間長度 / 古箏定位顏色」欄的顯示文字。"""
    dur   = duration_map.get(song, "")
    color = SONG_COLOR_MAP.get(song, "")
    if dur and color:
        return f"{dur}\n{color}"
    return color or dur


# =============================================================
# 列高自動調整
# =============================================================

def _auto_row_height(ws):
    for row in range(1, ws.max_row + 1):
        max_lines = max(
            (str(ws.cell(row, col).value or "").count("\n") + 1)
            for col in range(1, ws.max_column + 1)
        )
        ws.row_dimensions[row].height = max(22, 18 * max_lines)


# =============================================================
# 主要組裝函式
# =============================================================

def build_output_workbook(
    song_order:          List[str],
    song_assets:         Dict[str, SongAsset],
    song_people:         Dict[str, SongPeople],
    backstage_roles:     Dict[str, str],
    transition_time_map: Dict[TransitionKey, str],
    song_time_map:       Dict[str, str],
    duration_map:        Dict[str, str],
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "表演順序"

    # ── 表頭 ──────────────────────────────────────────────────
    for col, header_text, width in _COLUMNS:
        ws.cell(1, col, header_text)
        ws.column_dimensions[get_column_letter(col)].width = width

    style_row(ws, 1, NUM_COLS, fill=header_fill(), bold=True)
    # 表頭字體大小單獨設定（比資料列大）
    for col in range(1, NUM_COLS + 1):
        ws.cell(1, col).font = ws.cell(1, col).font.copy(size=HEADER_FONT_SIZE)

    ws.freeze_panes = "C2"   # 凍結第一列 + 前兩欄

    current_row = 2

    # ── 內部工具：寫入一個換場列 ──────────────────────────────
    def write_transition_row(
        prev_song:           Optional[str],
        next_song:           Optional[str],
        row_type:            str,
        is_pre:              bool = False,
        before_intermission: bool = False,
        last_teardown:       bool = False,
    ):
        nonlocal current_row

        left, mid, right = generate_transition(
            prev_song, next_song,
            song_assets.get(prev_song) if prev_song else None,
            song_assets.get(next_song) if next_song else None,
            song_people.get(prev_song) if prev_song else None,
            song_people.get(next_song) if next_song else None,
            backstage_roles,
            is_pre_row=is_pre,
            is_before_intermission=before_intermission,
            is_last_song_teardown=last_teardown,
        )

        ws.cell(current_row, 1, transition_time_map.get((prev_song, next_song, row_type), ""))
        ws.cell(current_row, 5, left)
        ws.cell(current_row, 6, mid)
        ws.cell(current_row, 7, right)
        style_row(ws, current_row, NUM_COLS, fill=None)
        current_row += 1

    # ── 開場前置換場列 ────────────────────────────────────────
    if song_order:
        write_transition_row(None, song_order[0], "pre_opening", is_pre=True)

    # ── 逐首寫入 ──────────────────────────────────────────────
    for idx, song in enumerate(song_order):
        is_last = idx == len(song_order) - 1

        # 曲目列
        ws.cell(current_row, 1, song_time_map.get(song, ""))
        ws.cell(current_row, 2, song)
        ws.cell(current_row, 3, _duration_color_text(song, duration_map))
        ws.cell(current_row, 4, _format_performer_cell(song_people[song]))
        style_row(ws, current_row, NUM_COLS, fill=song_fill())
        current_row += 1

        # 中場休息段落
        if INTERMISSION_AFTER is not None and idx + 1 == INTERMISSION_AFTER and not is_last:
            next_song = song_order[idx + 1]

            write_transition_row(song, next_song, "before_intermission", before_intermission=True)

            ws.cell(current_row, 1, transition_time_map.get((song, next_song, "intermission"), ""))
            ws.cell(current_row, 2, INTERMISSION_LABEL)
            style_row(ws, current_row, NUM_COLS, fill=intermission_fill())
            current_row += 1

            write_transition_row(None, next_song, "pre_second_half", is_pre=True)
            continue

        # 一般換場列 / 最後一首拆台列
        next_song = song_order[idx + 1] if not is_last else None
        write_transition_row(
            song, next_song,
            row_type="last_teardown" if is_last else "normal_transition",
            last_teardown=is_last,
        )

    # ── 結尾區塊 ──────────────────────────────────────────────
    ws.cell(current_row, 2, ENDING_BLOCK)
    border = thin_border()
    for col in range(1, NUM_COLS + 1):
        fill = intermission_fill() if col == 2 else None
        style_cell(ws.cell(current_row, col), fill=fill)
        ws.cell(current_row, col).border = border

    _auto_row_height(ws)
    return wb
