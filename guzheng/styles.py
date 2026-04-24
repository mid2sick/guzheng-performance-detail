# -*- coding: utf-8 -*-
"""
Excel 視覺樣式層。
只負責顏色、字體、框線等純呈現設定，不包含任何業務邏輯。
"""
from __future__ import annotations
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook

# =========================
# 色票
# =========================
_COLOR_HEADER       = "D9E2F3"   # 表頭列：淡藍
_COLOR_SONG         = "DDEBF7"   # 曲目列：更淡藍
_COLOR_INTERMISSION = "EDEDED"   # 中場 / 結尾：淺灰
_COLOR_BORDER       = "D9D9D9"   # 細框線：淺灰

# 預設儲存格字體（資料列）
_DEFAULT_CELL_FONT = "Microsoft JhengHei"
_DEFAULT_CELL_SIZE = 12

# 表頭字體大小
HEADER_FONT_SIZE = 14


# =========================
# 可重複使用的 Style 物件
# =========================
def header_fill()       -> PatternFill: return PatternFill("solid", fgColor=_COLOR_HEADER)
def song_fill()         -> PatternFill: return PatternFill("solid", fgColor=_COLOR_SONG)
def intermission_fill() -> PatternFill: return PatternFill("solid", fgColor=_COLOR_INTERMISSION)
def thin_border()       -> Border:      return Border(bottom=Side(style="thin", color=_COLOR_BORDER))


# =========================
# 儲存格樣式套用
# =========================
def style_cell(
    cell:      Cell,
    fill:      PatternFill | None = None,
    bold:      bool               = False,
    size:      int                = _DEFAULT_CELL_SIZE,
    font_name: str                = _DEFAULT_CELL_FONT,
):
    """套用對齊、字體、填色到單一儲存格。"""
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.font = Font(name=font_name, size=size, bold=bold)
    if fill is not None:
        cell.fill = fill


def style_row(ws, row: int, num_cols: int, fill: PatternFill | None, bold: bool = False):
    """對整列所有欄套用相同樣式與細框線。"""
    border = thin_border()
    for col in range(1, num_cols + 1):
        cell = ws.cell(row, col)
        style_cell(cell, fill=fill, bold=bold)
        cell.border = border


# =========================
# 全域字體大小統一
# =========================
def apply_global_font_size(wb: Workbook, size: int = 10, default_name: str = "Calibri"):
    """
    保留原本粗體 / 斜體 / 底線等設定，只統一字體大小為指定值。
    通常在 workbook 完成後最後套用一次。
    """
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                f = cell.font
                cell.font = Font(
                    name      = f.name      if f and f.name      else default_name,
                    size      = size,
                    bold      = f.bold      if f else False,
                    italic    = f.italic    if f else False,
                    underline = f.underline if f else None,
                    strike    = f.strike    if f else False,
                    color     = f.color     if f else None,
                    vertAlign = f.vertAlign if f else None,
                    outline   = f.outline   if f else False,
                    shadow    = f.shadow    if f else False,
                    charset   = f.charset   if f else None,
                    scheme    = f.scheme    if f else None,
                    family    = f.family    if f else None,
                    condense  = f.condense  if f else False,
                    extend    = f.extend    if f else False,
                )
